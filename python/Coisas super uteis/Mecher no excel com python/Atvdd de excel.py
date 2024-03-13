import openpyxl

import openpyxl

# Criar uma planilha (Pasta)
book = openpyxl.Workbook()

# Como visualizar páginas existentes
print(book.sheetnames)

# Como criar uma página
book.create_sheet('Computadores')

# Como escolher uma página
frutas_page = book['Computadores']

# Escrever algo
frutas_page.append(['Eletetrônico', 'Quantidades', 'Preço'])
frutas_page.append(['pc1', '8Gb ram', 'R$3090'])
frutas_page.append(['pc2', '16Gb ram', 'R$4000'])
frutas_page.append(['pc3', '32Gb ram', 'R$6000'])

# Salvar planilha
book.save('meus computadores.xlsx')
