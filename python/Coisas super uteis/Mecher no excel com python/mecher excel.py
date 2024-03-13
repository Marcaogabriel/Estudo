import openpyxl

# Criar uma planilha (Pasta)
book = openpyxl.Workbook()

# Como visualizar páginas existentes
print(book.sheetnames)

# Como criar uma página
book.create_sheet('Frutas')

# Como escolher uma página
frutas_page = book['Frutas']

# Escrever algo
frutas_page.append(['Frutas', 'Quantidades', 'Preço'])
frutas_page.append(['Banana', '5', 'R$3,90'])
frutas_page.append(['maça', '3', 'R$4,00'])
frutas_page.append(['Mamão', '1', 'R$6,00'])
frutas_page.append(['Limão', '5', 'R$2,00'])

# Salvar planilha
book.save('Planilha de compras.xlsx')
